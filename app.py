# Absolutely screwed the pooch on this. I did not realize that Render free tier restarts your service every 15 minutes...
# Self to self, if you for some reason are ever going to do something like this again, make sure to save the log file to Github or even rewrite using React/JS and use Vercels hosting and their integrated DB.
# To be honest, I hid the timestamps in the dashboard to make it less obvious that all the data before 4:00pm was wiped.
# Kind of snitching on myself but I had to admit it somewhere.
# Don't use Render without persistent storage!!!

from flask import Flask, render_template, request, redirect, send_file
import os
import csv
import pandas as pd
from datetime import datetime

app = Flask(__name__)

LOG_FILE = "game_log.csv"
SECRET_DELETE_CODE = "DELETE4220"

CATEGORIES = [
    "Civil Rights and Duties",
    "Historical Migration",
    "Local Migration",
    "Borders",
    "Migrant Experiences / Myths",
    "Statistics / Miscellaneous"
]

DIFFICULTIES = ["Easy", "Medium", "Hard"]
YEARS = ["Freshman", "Sophomore", "Junior", "Senior", "Grad", "Faculty", "Public"]

# =====================================================
# CREATE LOG FILE IF NEEDED
# =====================================================
def ensure_logfile():
    if not os.path.exists(LOG_FILE):
        with open(LOG_FILE, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerow(["timestamp", "category", "difficulty", "correct", "year", "participant_id"])
ensure_logfile()

# =====================================================
# READ LOG
# =====================================================
def read_logfile(path):
    EXPECTED = ["timestamp", "category", "difficulty", "correct", "year", "participant_id"]

    if not os.path.exists(path) or os.path.getsize(path) == 0:
        return pd.DataFrame(columns=EXPECTED)

    df = pd.read_csv(path)

    for col in EXPECTED:
        if col not in df.columns:
            df[col] = None

    df["participant_id"] = pd.to_numeric(df["participant_id"], errors="coerce").fillna(-1).astype(int)

    df["year"] = (
        df["year"]
        .replace(["", "None", None], pd.NA)
        .groupby(df["participant_id"])
        .transform(lambda x: x.bfill().ffill())
    )

    return df

# =====================================================
# LAST PARTICIPANT
# =====================================================
def get_last_participant_id():
    df = read_logfile(LOG_FILE)

    # No entries → start at 1 always
    if df.empty:
        return 1

    # If no valid participant IDs, also start at 1
    valid_ids = df["participant_id"].dropna().astype(int)
    if valid_ids.empty:
        return 1

    return int(valid_ids.max())

# =====================================================
# HOME
# =====================================================
@app.route("/")
def home():
    return redirect("/log")

@app.route("/log", methods=["GET", "POST"])
def log():
    # -----------------------------------------
    # GET = Load form initially
    # -----------------------------------------
    if request.method == "GET":
        step = int(request.args.get("step", 1))
        step = min(max(step, 1), 5)

        last_pid = get_last_participant_id()
        active_pid = last_pid if last_pid > 0 else None

        return render_template(
            "log.html",
            categories=CATEGORIES,
            difficulties=DIFFICULTIES,
            years=YEARS,
            step=step,
            prev_category=None,
            prev_difficulty=None,
            prev_correct=None,
            prev_year=None,
            participant_id=active_pid,
            skip_year="0",          # start with no skipping
            success=False,
            error=None
        )

    # -----------------------------------------
    # POST = Move through steps or submit
    # -----------------------------------------
    form = request.form
    step = int(form.get("step", 1))
    action = form.get("action", "next")
    skip_year = form.get("skip_year", "0")

    # Values carried between steps
    prev_category = form.get("category") or None
    prev_difficulty = form.get("difficulty") or None
    prev_correct = form.get("correct") or None
    prev_year = form.get("year") or None
    prev_pid = form.get("participant_id")

    # Validate participant ID
    if prev_pid and prev_pid.isdigit():
        prev_pid = int(prev_pid)
    else:
        prev_pid = get_last_participant_id()
        if prev_pid == 0:
            prev_pid = 1

    # -----------------------------------------
    # UPDATE previous values from current step
    # -----------------------------------------
    if form.get("category_new"):
        prev_category = form.get("category_new")
    if form.get("difficulty_new"):
        prev_difficulty = form.get("difficulty_new")
    if form.get("correct_new"):
        prev_correct = form.get("correct_new")
    if form.get("year_new"):
        prev_year = form.get("year_new")

    # -----------------------------------------
    # BACK BUTTON: just go back one step
    # -----------------------------------------
    if action == "back":
        return render_template(
            "log.html",
            categories=CATEGORIES,
            difficulties=DIFFICULTIES,
            years=YEARS,
            step=max(1, step - 1),
            prev_category=prev_category,
            prev_difficulty=prev_difficulty,
            prev_correct=prev_correct,
            prev_year=prev_year,
            participant_id=prev_pid,
            skip_year=skip_year,
            success=False,
            error=None
        )

    # -----------------------------------------
    # VALIDATION (AFTER updating)
    # -----------------------------------------
    error = None
    if step == 1 and not prev_category:
        error = "Pick a category."
    elif step == 2 and not prev_difficulty:
        error = "Pick a difficulty."
    elif step == 3 and prev_correct not in ("0", "1"):
        error = "Select correct/incorrect."
    elif step == 4 and not prev_year:
        error = "Pick a year."

    # If there is an error, stay on the same step
    if error:
        return render_template(
            "log.html",
            categories=CATEGORIES,
            difficulties=DIFFICULTIES,
            years=YEARS,
            step=step,
            prev_category=prev_category,
            prev_difficulty=prev_difficulty,
            prev_correct=prev_correct,
            prev_year=prev_year,
            participant_id=prev_pid,
            skip_year=skip_year,
            success=False,
            error=error
        )

    # -----------------------------------------
    # STEP 5 → FINAL SUBMIT
    # -----------------------------------------
    if step == 5 and action == "submit":
        df = read_logfile(LOG_FILE)

        # Participant's last year (if exists)
        if not df.empty and prev_pid in df["participant_id"].values:
            last_year = df[df["participant_id"] == prev_pid]["year"].dropna().iloc[-1]
        else:
            last_year = None

        if prev_year is None:
            prev_year = last_year or "Unknown"

        # Participant choice for next round
        user_choice = form.get("participant_choice", "same")

        if user_choice == "new":
            pid_for_entry = prev_pid + 1
            skip_year_next = "0"   # new participant → must choose year next time
        else:
            pid_for_entry = prev_pid
            skip_year_next = "1"   # same participant → skip year next time

        # Write entry
        with open(LOG_FILE, "a", newline="") as f:
            writer = csv.writer(f)
            writer.writerow([
                datetime.now().isoformat(),
                prev_category,
                prev_difficulty,
                "Yes" if prev_correct == "1" else "No",
                prev_year,
                pid_for_entry
            ])


        # Back to step 1 for next question
        return render_template(
            "log.html",
            categories=CATEGORIES,
            difficulties=DIFFICULTIES,
            years=YEARS,
            step=1,
            prev_category=None,
            prev_difficulty=None,
            prev_correct=None,
            prev_year=prev_year,
            participant_id=str(pid_for_entry),
            skip_year=skip_year_next,
            success=True,
            error=None
        )

    # -----------------------------------------
    # STEP 4 → move to REVIEW (step 5)
    # -----------------------------------------
    if step == 4 and action == "next":
        # We already validated year above, so just go to step 5
        return render_template(
            "log.html",
            categories=CATEGORIES,
            difficulties=DIFFICULTIES,
            years=YEARS,
            step=5,
            prev_category=prev_category,
            prev_difficulty=prev_difficulty,
            prev_correct=prev_correct,
            prev_year=prev_year,
            participant_id=prev_pid,
            skip_year=skip_year,
            success=False,
            error=None
        )

    # -----------------------------------------
    # OTHER STEPS: move forward, maybe skipping 4
    # -----------------------------------------
    next_step = min(step + 1, 5)

    # If the next step *would* be 4 but we're reusing the same participant,
    # skip directly to review (5).
    if next_step == 4 and skip_year == "1":
        next_step = 5

    return render_template(
        "log.html",
        categories=CATEGORIES,
        difficulties=DIFFICULTIES,
        years=YEARS,
        step=next_step,
        prev_category=prev_category,
        prev_difficulty=prev_difficulty,
        prev_correct=prev_correct,
        prev_year=prev_year,
        participant_id=prev_pid,
        skip_year=skip_year,
        success=False,
        error=None
    )

# =====================================================
# DASHBOARD
# =====================================================
@app.route("/dashboard")
def dashboard():
    df = read_logfile(LOG_FILE)

    if df.empty:
        return render_template(
            "dashboard.html",
            empty=True,
            grouped_entries=[],
            years=["All"] + YEARS,
            categories=["All"] + CATEGORIES,
            selected_year="All",
            selected_category="All",
            total=0,
            accuracy=0
        )

    selected_year = request.args.get("year", "All")
    selected_cat = request.args.get("category", "All")

    if selected_year != "All":
        df = df[df["year"] == selected_year]
    if selected_cat != "All":
        df = df[df["category"] == selected_cat]

    total = len(df)

    # Convert Yes/No into numbers only for accuracy %
    df["correct_num"] = df["correct"].map({"Yes": 1, "No": 0})
    valid_correct = df["correct_num"].dropna()
    accuracy = round(valid_correct.mean() * 100, 2) if not valid_correct.empty else 0

    grouped_entries = []
    for pid, sub in df.groupby("participant_id"):
        grouped_entries.append({
            "participant_id": pid,
            "rows": sub.assign(
                correct_text=lambda x: x["correct"].map({
                    "Yes": "Correct",
                    "No": "Incorrect"
                })

            )[
                ["timestamp", "year", "category", "difficulty", "correct_text"]
            ].to_dict(orient="records")
        })

    return render_template(
        "dashboard.html",
        empty=False,
        grouped_entries=grouped_entries,
        years=["All"] + YEARS,
        categories=["All"] + CATEGORIES,
        selected_year=selected_year,
        selected_category=selected_cat,
        total=total,
        accuracy=accuracy
    )

# =====================================================
# EXPORT CSV / EXCEL
# =====================================================
@app.route("/export/csv")
def export_csv():
    return send_file(LOG_FILE, as_attachment=True, download_name="migration_game_log.csv")

@app.route("/export/excel")
def export_excel():
    df = read_logfile(LOG_FILE)
    filename = "migration_game_log.xlsx"
    df.to_excel(filename, index=False)

    # ----- Formatting -----
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Border, Side

    wb = load_workbook(filename)
    ws = wb.active

    # Auto-size columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Header styling
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    # Borders
    thin = Side(border_style="thin", color="000000")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)
    for row in ws.iter_rows():
        for cell in row:
            cell.border = border

    wb.save(filename)

    return send_file(filename, as_attachment=True)


# =====================================================
# DELETE
# =====================================================
@app.route("/delete", methods=["GET", "POST"])
def delete_data():
    if request.method == "POST":
        code = request.form.get("confirm_text", "").strip()

        if code == SECRET_DELETE_CODE:
            with open(LOG_FILE, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(["timestamp", "category", "difficulty", "correct", "year", "participant_id"])
            return render_template("delete.html", deleted=True)

        return render_template("delete.html", error="Incorrect code.", deleted=False)

    return render_template("delete.html", deleted=False)

# =====================================================
# RUN
# =====================================================
if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)